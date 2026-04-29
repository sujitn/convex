"""Regenerate excel/ConvexDemo.xlsx from scratch against the current CX.* surface.

The workbook is a derived artifact — this script is the source of truth.
Run it whenever the cell API changes:

    pip install openpyxl
    python excel/build_demo.py

Open the resulting workbook with the Convex add-in loaded; every formula
should evaluate without #NAME? / #VALUE! errors.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT = Path(__file__).parent / "ConvexDemo.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color="1F4E78")
NOTE_FONT = Font(italic=True, color="595959")


def header(ws: Worksheet, row: int, cols: list[str]) -> None:
    for j, label in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=j, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def section(ws: Worksheet, row: int, label: str, span: int = 1) -> None:
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def note(ws: Worksheet, row: int, text: str, span: int = 6) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def fit_columns(ws: Worksheet, widths: list[int]) -> None:
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ---------------------------------------------------------------------------
# README sheet
# ---------------------------------------------------------------------------

def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    fit_columns(ws, [120])

    lines = [
        "Convex Excel Add-In — Demo workbook",
        "",
        "Every formula here uses the CX.* UDF surface defined in excel/README.md.",
        "Open with the Convex add-in loaded (the packed .xll under bin/Release/net472/publish/).",
        "",
        "Sheets:",
        "  • Bonds      — build fixed/callable/FRN/zero, mark-driven price + risk grid.",
        "  • Curves     — discrete + bootstrapped curves, plot points via CX.CURVE.QUERY.",
        "  • Spreads    — Z / I / ASW on a swap curve; G-spread on a separate gov curve.",
        "  • Scenarios  — parallel-shift bps ladder against the base bond.",
        "  • Schemas    — JSON wire format for any DTO via CX.SCHEMA.",
        "",
        "Trader-mark grammar (forwarded to the Rust parser):",
        "  99.5            clean price (default)",
        "  99.5C / 99.5D   clean / dirty",
        "  99-16+          Treasury 32nds (+ = ½)",
        "  4.65%@SA        yield + frequency",
        "  +125bps@USD.SOFR   Z-spread (default) over benchmark",
        "  125 OAS@USD.TSY    explicit spread family",
        "",
        "Conventions:",
        "  • Coupon rates: decimal (0.05 = 5%).",
        "  • Spread bps:   bps (75 = 75 bp).",
        "  • OAS volatility: decimal (0.01 = 1%).",
        "  • Prices:       per 100 face.",
        "",
        "If a formula returns #NAME? the add-in is not loaded.",
        "If it returns a text string starting #ERROR: ... that's a structured FFI error;",
        "the message tells you which argument was rejected.",
    ]
    for i, line in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="1F4E78")


# ---------------------------------------------------------------------------
# Bonds sheet
# ---------------------------------------------------------------------------

def build_bonds(wb: Workbook) -> None:
    ws = wb.create_sheet("Bonds")
    fit_columns(ws, [22, 16, 10, 14, 14, 16, 16, 14, 14, 14, 14])

    section(ws, 1, "Build bonds with =CX.BOND / .CALLABLE / .FRN / .ZERO", span=11)

    # Inputs table
    header(ws, 3, [
        "Identifier", "Type", "Coupon", "Maturity", "Issue", "Frequency",
        "Day count", "Handle",
    ])

    today_iso = date.today().isoformat()
    inputs = [
        ("FIXED-5Y",  "fixed_rate",   0.045,  date(2030, 1, 15), date(2025, 1, 15), "SemiAnnual", "Thirty360US"),
        ("FIXED-10Y", "fixed_rate",   0.05,   date(2035, 1, 15), date(2025, 1, 15), "SemiAnnual", "Thirty360US"),
        ("FIXED-30Y", "fixed_rate",   0.052,  date(2055, 1, 15), date(2025, 1, 15), "SemiAnnual", "Thirty360US"),
    ]

    for i, (cusip, _kind, coupon, maturity, issue, freq, dc) in enumerate(inputs, start=4):
        ws.cell(row=i, column=1, value=cusip)
        ws.cell(row=i, column=2, value="fixed_rate")
        ws.cell(row=i, column=3, value=coupon)
        ws.cell(row=i, column=4, value=maturity)
        ws.cell(row=i, column=5, value=issue)
        ws.cell(row=i, column=6, value=freq)
        ws.cell(row=i, column=7, value=dc)
        ws.cell(row=i, column=8, value=(
            f'=CX.BOND(A{i},C{i},D{i},E{i},F{i},G{i})'
        ))

    # FRN row
    frn_row = 7
    ws.cell(row=frn_row, column=1, value="FRN-5Y-SOFR")
    ws.cell(row=frn_row, column=2, value="floating_rate")
    ws.cell(row=frn_row, column=3, value=75)  # spread bps
    ws.cell(row=frn_row, column=4, value=date(2030, 1, 15))
    ws.cell(row=frn_row, column=5, value=date(2025, 1, 15))
    ws.cell(row=frn_row, column=6, value="Quarterly")
    ws.cell(row=frn_row, column=7, value="Act360")
    ws.cell(row=frn_row, column=8, value=(
        f'=CX.BOND.FRN(A{frn_row},C{frn_row},D{frn_row},E{frn_row},"sofr",F{frn_row},G{frn_row})'
    ))

    # Zero row
    zero_row = 8
    ws.cell(row=zero_row, column=1, value="ZCB-5Y")
    ws.cell(row=zero_row, column=2, value="zero_coupon")
    ws.cell(row=zero_row, column=3, value="—")
    ws.cell(row=zero_row, column=4, value=date(2030, 1, 15))
    ws.cell(row=zero_row, column=5, value=date(2025, 1, 15))
    ws.cell(row=zero_row, column=6, value="SemiAnnual")
    ws.cell(row=zero_row, column=7, value="ActActIcma")
    ws.cell(row=zero_row, column=8, value=(
        f'=CX.BOND.ZERO(A{zero_row},D{zero_row},E{zero_row},F{zero_row},G{zero_row})'
    ))

    # Callable row. Excel rejects array constants that contain function calls
    # (e.g. `{DATE(2030,1,15);...}`), so the schedule lives in a real range
    # the user can edit; CX.BOND.CALLABLE consumes it as parallel ranges.
    call_row = 9
    ws.cell(row=call_row, column=1, value="CALL-10NC5")
    ws.cell(row=call_row, column=2, value="callable")
    ws.cell(row=call_row, column=3, value=0.06)
    ws.cell(row=call_row, column=4, value=date(2035, 1, 15))
    ws.cell(row=call_row, column=5, value=date(2025, 1, 15))
    ws.cell(row=call_row, column=6, value="SemiAnnual")
    ws.cell(row=call_row, column=7, value="Thirty360US")

    # Schedule table starts two rows below the callable row.
    sched_header_row = call_row + 2
    ws.cell(row=sched_header_row, column=9,  value="Call date").font = Font(bold=True)
    ws.cell(row=sched_header_row, column=10, value="Call price").font = Font(bold=True)
    schedule = [
        (date(2030, 1, 15), 102.0),
        (date(2032, 1, 15), 101.0),
        (date(2034, 1, 15), 100.0),
    ]
    sched_first = sched_header_row + 1
    for k, (d, p) in enumerate(schedule):
        ws.cell(row=sched_first + k, column=9,  value=d)
        ws.cell(row=sched_first + k, column=10, value=p)
    sched_last = sched_first + len(schedule) - 1

    ws.cell(row=call_row, column=8, value=(
        f'=CX.BOND.CALLABLE(A{call_row},C{call_row},D{call_row},E{call_row},'
        f'I{sched_first}:I{sched_last},J{sched_first}:J{sched_last},'
        f'F{call_row},"american",G{call_row})'
    ))

    # Pricing & risk grid (offset down to clear the schedule table).
    grid_section_row = sched_last + 2
    grid_header_row = grid_section_row + 2
    section(
        ws,
        grid_section_row,
        "Mark-driven pricing — change the mark and watch every column recompute",
        span=11,
    )
    header(ws, grid_header_row, [
        "Bond", "Settlement", "Mark", "Clean", "Dirty", "Accrued",
        "YTM (%)", "Mod Dur", "Mac Dur", "Convexity", "DV01",
    ])

    # Reference the handle cells from the inputs table.
    grid_first = grid_header_row + 1
    for offset, src_row in enumerate([4, 5, 6]):  # the three fixed-rate bonds
        r = grid_first + offset
        ws.cell(row=r, column=1, value=f'=A{src_row}')
        ws.cell(row=r, column=2, value=date.today())
        ws.cell(row=r, column=3, value="99.5C")
        ws.cell(row=r, column=4, value=f'=CX.PRICE(H{src_row},B{r},C{r})')
        ws.cell(row=r, column=5, value=f'=CX.PRICE(H{src_row},B{r},C{r},,,"dirty")')
        ws.cell(row=r, column=6, value=f'=CX.PRICE(H{src_row},B{r},C{r},,,"accrued")')
        ws.cell(row=r, column=7, value=f'=CX.PRICE(H{src_row},B{r},C{r},,,"ytm")')
        ws.cell(row=r, column=8, value=f'=CX.RISK(H{src_row},B{r},C{r},,"mod_dur")')
        ws.cell(row=r, column=9, value=f'=CX.RISK(H{src_row},B{r},C{r},,"mac_dur")')
        ws.cell(row=r, column=10, value=f'=CX.RISK(H{src_row},B{r},C{r},,"convexity")')
        ws.cell(row=r, column=11, value=f'=CX.RISK(H{src_row},B{r},C{r},,"dv01")')

    grid_last = grid_first + 2
    note_row = grid_last + 1
    note(ws, note_row,
         f"Try changing the mark in C{grid_first}:C{grid_last} — "
         "'99.5C', '4.65%@SA', or '99-16+' all work. "
         "The grid above recomputes each column.")

    cashflow_section_row = note_row + 2
    section(ws, cashflow_section_row, "Cashflows — =CX.CASHFLOWS(bond, settle)", span=11)
    cashflow_formula_row = cashflow_section_row + 1
    ws.cell(row=cashflow_formula_row, column=1,
            value=f'=CX.CASHFLOWS(H5,B{grid_first + 1})')
    note(ws, cashflow_formula_row + 1,
         "Spills a 3-column grid: date, amount per 100 face, kind "
         "(coupon | redemption | coupon-and-redemption | fee).")


# ---------------------------------------------------------------------------
# Curves sheet
# ---------------------------------------------------------------------------

def build_curves(wb: Workbook) -> None:
    ws = wb.create_sheet("Curves")
    fit_columns(ws, [16, 14, 14, 16, 16, 16])

    section(ws, 1, "Discrete curve — =CX.CURVE", span=6)
    header(ws, 3, ["Tenor (yrs)", "Zero rate (decimal)"])
    pillars = [
        (0.25, 0.0525), (0.5, 0.0530), (1.0, 0.0500),
        (2.0, 0.0470), (5.0, 0.0445), (10.0, 0.0440),
        (30.0, 0.0455),
    ]
    for i, (t, r) in enumerate(pillars, start=4):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=r)
    last = 3 + len(pillars)

    ws.cell(row=last + 2, column=1, value="Curve handle:")
    ws.cell(row=last + 2, column=2, value=(
        f'=CX.CURVE("USD.SOFR.DEMO",TODAY(),A4:A{last},B4:B{last},'
        f'"zero_rate","linear","Act365Fixed","Continuous")'
    ))

    section(ws, last + 4, "Bootstrap — =CX.CURVE.BOOTSTRAP", span=6)
    header(ws, last + 6, ["Kind", "Tenor", "Rate (decimal)"])
    instruments = [
        ("deposit", 0.25, 0.0525),
        ("deposit", 0.5,  0.0530),
        ("swap",    2.0,  0.0470),
        ("swap",    5.0,  0.0445),
        ("swap",    10.0, 0.0440),
        ("swap",    30.0, 0.0455),
    ]
    base = last + 7
    for i, (k, t, r) in enumerate(instruments, start=base):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=t)
        ws.cell(row=i, column=3, value=r)
    inst_end = base + len(instruments) - 1

    ws.cell(row=inst_end + 2, column=1, value="Bootstrapped:")
    ws.cell(row=inst_end + 2, column=2, value=(
        f'=CX.CURVE.BOOTSTRAP("USD.SOFR.BOOT",TODAY(),'
        f'A{base}:A{inst_end},B{base}:B{inst_end},C{base}:C{inst_end},'
        f'"global_fit","linear","Act360")'
    ))

    section(ws, inst_end + 4, "Read points off the curve — =CX.CURVE.QUERY", span=6)
    header(ws, inst_end + 6, ["Tenor", "Zero (%)", "DF", "1Y forward (%)"])
    sweep_start = inst_end + 7
    handle_cell = f'B{last + 2}'  # discrete curve handle
    for i, t in enumerate([0.25, 0.5, 1, 2, 3, 5, 7, 10, 15, 20, 30], start=sweep_start):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=f'=CX.CURVE.QUERY({handle_cell},A{i},"zero")*100')
        ws.cell(row=i, column=3, value=f'=CX.CURVE.QUERY({handle_cell},A{i},"df")')
        ws.cell(row=i, column=4, value=f'=CX.CURVE.QUERY({handle_cell},A{i},"forward",A{i}+1)*100')

    note(ws, sweep_start + 12,
         "Plot column B against column A for the zero curve, "
         "column D for the 1Y forward curve overlay.")


# ---------------------------------------------------------------------------
# Spreads sheet
# ---------------------------------------------------------------------------

def build_spreads(wb: Workbook) -> None:
    ws = wb.create_sheet("Spreads")
    fit_columns(ws, [22, 22, 14, 14, 14, 14, 14])

    section(ws, 1, "Bond × discount curve × mark → spreads", span=7)
    note(ws, 2,
         "Reuses the FIXED-10Y handle from the Bonds sheet and the discrete curve from Curves. "
         "G-spread requires its own government curve — built below.")

    ws.cell(row=4, column=1, value="Bond (10y handle):")
    ws.cell(row=4, column=2, value="=Bonds!H5")
    ws.cell(row=5, column=1, value="Discount curve (swap):")
    ws.cell(row=5, column=2, value="=Curves!B12")  # the discrete curve handle
    ws.cell(row=6, column=1, value="Settlement:")
    ws.cell(row=6, column=2, value=date.today())
    ws.cell(row=7, column=1, value="Mark:")
    ws.cell(row=7, column=2, value="99.5C")

    header(ws, 9, ["Spread", "bps", "Spread DV01", "Notes"])
    ws.cell(row=10, column=1, value="Z-spread")
    ws.cell(row=10, column=2, value='=CX.SPREAD(B4,B5,B6,B7,"Z")')
    ws.cell(row=10, column=3, value='=INDEX(CX.SPREAD(B4,B5,B6,B7,"Z",,"grid"),2,2)')
    ws.cell(row=10, column=4, value="Constant spread over the spot discount curve.")

    ws.cell(row=11, column=1, value="I-spread")
    ws.cell(row=11, column=2, value='=CX.SPREAD(B4,B5,B6,B7,"I")')
    ws.cell(row=11, column=4, value="Yield − interpolated swap rate at maturity.")

    ws.cell(row=12, column=1, value="ASW (par)")
    ws.cell(row=12, column=2, value='=CX.SPREAD(B4,B5,B6,B7,"ASW")')
    ws.cell(row=12, column=4, value="Par-par asset swap spread.")

    ws.cell(row=13, column=1, value="ASW (proceeds)")
    ws.cell(row=13, column=2, value='=CX.SPREAD(B4,B5,B6,B7,"ASW_PROC")')
    ws.cell(row=13, column=4, value="Proceeds asset swap spread.")

    section(ws, 15, "G-spread requires a separate government curve", span=7)
    header(ws, 17, ["Tenor", "Govt zero (decimal)"])
    govt_pillars = [(2, 0.042), (5, 0.044), (10, 0.045), (30, 0.046)]
    for i, (t, r) in enumerate(govt_pillars, start=18):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=r)
    govt_end = 17 + len(govt_pillars)

    ws.cell(row=govt_end + 2, column=1, value="Govt curve handle:")
    ws.cell(row=govt_end + 2, column=2, value=(
        f'=CX.CURVE("USD.TSY.DEMO",TODAY(),A18:A{govt_end},B18:B{govt_end},'
        f'"zero_rate","linear","Act365Fixed","Continuous")'
    ))

    note(ws, govt_end + 4,
         "G-spread today is wired only via the Spread Ticket ribbon form, which threads "
         "params.govt_curve through. The cell-side helper UDF will land in the next pass; "
         "open the ribbon Spread Ticket to compute G-spread interactively.")


# ---------------------------------------------------------------------------
# Scenarios sheet
# ---------------------------------------------------------------------------

def build_scenarios(wb: Workbook) -> None:
    ws = wb.create_sheet("Scenarios")
    fit_columns(ws, [16, 16, 16, 16, 16, 18])

    section(ws, 1, "Parallel-shift scenarios via mark text", span=6)
    note(ws, 2,
         "Hold the bond fixed; bump the YTM by N bps and reprice. Each row computes the "
         "yield mark as <base + Δ>%@SA — the same path the ScenarioForm uses internally.")

    ws.cell(row=4, column=1, value="Bond (10y handle):")
    ws.cell(row=4, column=2, value="=Bonds!H5")
    ws.cell(row=5, column=1, value="Settlement:")
    ws.cell(row=5, column=2, value=date.today())
    ws.cell(row=6, column=1, value="Base mark:")
    ws.cell(row=6, column=2, value="99.5C")
    ws.cell(row=7, column=1, value="Base YTM (%):")
    ws.cell(row=7, column=2, value='=CX.PRICE(B4,B5,B6,,,"ytm")')
    ws.cell(row=8, column=1, value="Base clean:")
    ws.cell(row=8, column=2, value='=CX.PRICE(B4,B5,B6)')

    header(ws, 10, ["Shift (bps)", "Bumped yield (%)", "Bumped mark", "Clean", "ΔP (clean)"])
    shifts = [-100, -50, -25, -10, 0, 10, 25, 50, 100]
    for i, sh in enumerate(shifts, start=11):
        ws.cell(row=i, column=1, value=sh)
        ws.cell(row=i, column=2, value=f'=$B$7+A{i}/100')
        ws.cell(row=i, column=3, value=f'=TEXT(B{i},"0.00000000")&"%@SA"')
        ws.cell(row=i, column=4, value=f'=CX.PRICE($B$4,$B$5,C{i})')
        ws.cell(row=i, column=5, value=f'=D{i}-$B$8')


# ---------------------------------------------------------------------------
# Schemas sheet
# ---------------------------------------------------------------------------

def build_schemas(wb: Workbook) -> None:
    ws = wb.create_sheet("Schemas")
    fit_columns(ws, [22, 110])

    section(ws, 1, "JSON wire-format schemas — =CX.SCHEMA(name)", span=2)
    note(ws, 2, "Each cell pulls the schema from the FFI directly. Fold a row to read.")

    types = [
        "Mark", "BondSpec", "CurveSpec",
        "PricingRequest", "PricingResponse",
        "RiskRequest", "RiskResponse",
        "SpreadRequest", "SpreadResponse",
        "CashflowRequest", "CashflowResponse",
        "CurveQueryRequest", "CurveQueryResponse",
    ]
    for i, t in enumerate(types, start=4):
        ws.cell(row=i, column=1, value=t).font = Font(bold=True)
        ws.cell(row=i, column=2, value=f'=CX.SCHEMA("{t}")').alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.row_dimensions[i].height = 110


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    wb = Workbook()
    build_readme(wb)
    build_bonds(wb)
    build_curves(wb)
    build_spreads(wb)
    build_scenarios(wb)
    build_schemas(wb)
    wb.save(OUTPUT)
    print(f"wrote {OUTPUT}")


if __name__ == "__main__":
    main()
