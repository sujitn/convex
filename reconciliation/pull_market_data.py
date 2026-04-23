"""
Pull primary-source data for the reconciliation harness.

US: FRED CMT, NY Fed SOFR, TreasuryDirect TIPS.
UK: BoE nominal spot curve zip (xlsx).
EU: ECB SDMX AAA euro-area spot yields.
JP: MOF Japan JGB historical par yields CSV.

All sources are free public endpoints, no API keys. Intermediate files are
written next to this script; curves.json stays hand-curated — refresh by
re-running this script, eyeballing the CSVs, and updating curves.json.

UK pull requires openpyxl (BoE publishes xlsx only). Other pulls are stdlib.

Run:
    pip install openpyxl       # optional, only for pull_uk_gilt
    python reconciliation/pull_market_data.py
"""
from __future__ import annotations

import csv
import io
import json
import math
import pathlib
import sys
import urllib.request
import urllib.error
import zipfile

HERE = pathlib.Path(__file__).parent
VAL_DATE = "2025-12-31"


def fetch(url: str, timeout: int = 60) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": "convex-reconciliation/0.1"})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read()
    except urllib.error.HTTPError as e:
        print(f"HTTP {e.code} fetching {url}: {e.reason}", file=sys.stderr)
        raise


def cont_to_sa(r_continuous_pct: float) -> float:
    """Continuous-compounded yield in percent → semi-annual equivalent in percent."""
    r = r_continuous_pct / 100.0
    return 2 * (math.exp(r / 2) - 1) * 100.0


def pull_ust_cmt() -> None:
    """Treasury constant-maturity yields on the valuation date from FRED."""
    series = ["DGS1MO", "DGS3MO", "DGS6MO", "DGS1", "DGS2",
              "DGS3", "DGS5", "DGS7", "DGS10", "DGS20", "DGS30"]
    url = (
        "https://fred.stlouisfed.org/graph/fredgraph.csv"
        f"?id={','.join(series)}&cosd={VAL_DATE}&coed={VAL_DATE}"
    )
    out = HERE / "ust_cmt_20251231.csv"
    out.write_bytes(fetch(url))
    print(f"wrote {out} ({out.stat().st_size} bytes)")


def pull_sofr_fixings() -> None:
    """Daily SOFR from NY Fed 2024-01-01 through 2025-12-31."""
    url = (
        "https://markets.newyorkfed.org/api/rates/secured/sofr/search.json"
        "?startDate=2024-01-01&endDate=2025-12-31"
    )
    raw = fetch(url).decode("utf-8")
    data = json.loads(raw)

    rows = data.get("refRates") if isinstance(data, dict) else data
    if rows is None:
        print("unexpected SOFR payload shape; dumping raw", file=sys.stderr)
        (HERE / "sofr_fixings.raw.json").write_text(raw)
        return

    out = HERE / "sofr_fixings.csv"
    with out.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["effective_date", "rate_pct", "volume_usd_bn"])
        for r in rows:
            w.writerow([
                r.get("effectiveDate"),
                r.get("percentRate"),
                r.get("volumeInBillions"),
            ])
    print(f"wrote {out} ({len(rows)} rows)")


def pull_tips_index_ratio() -> None:
    """TIPS 91282CNS6 index ratio on the valuation date (daily-published)."""
    cusip = "91282CNS6"
    url = (
        "https://www.treasurydirect.gov/TA_WS/securities/search"
        f"?cusip={cusip}&format=json"
    )
    raw = fetch(url).decode("utf-8")

    (HERE / "tips_search_raw.json").write_text(raw)
    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        print("tips payload was not json; raw saved", file=sys.stderr)
        return

    out = HERE / "tips_index_ratio_20251231.json"
    out.write_text(json.dumps({
        "cusip": cusip,
        "valuation_date": VAL_DATE,
        "index_ratio": None,
        "raw_search": data,
        "manual_pull_url": (
            "https://www.treasurydirect.gov/auctions/"
            "announcements-data-results/tips-cpi-data/tips-cpi-detail/"
            f"?cusip={cusip}"
        ),
        "note": (
            "The index-ratio time series lives on the TIPS/CPI detail page. "
            "If this script doesn't capture it automatically, open the URL "
            "above, download the CSV for the December 2025 range, and edit "
            "index_ratio manually."
        ),
    }, indent=2))
    print(f"wrote {out}")


# --------------------------------------------------------------- UK / EU / JP

# Tenors we want for reconciliation — keep a few past 10Y in case we add longer bonds.
CURVE_TENORS = [1, 2, 3, 5, 7, 10, 15, 20, 30]


def pull_uk_gilt() -> None:
    """BoE nominal spot curve on 2025-12-31.

    BoE publishes xlsx only (no CSV endpoint), so this needs openpyxl. The
    archive zip holds multi-year workbooks; we grab the 2025-to-present one,
    read sheet '4. spot curve', and emit a CSV with both continuous (BoE
    native) and semi-annual equivalent rates.
    """
    try:
        import openpyxl  # noqa: F401 — required below
    except ImportError:
        print("pull_uk_gilt: openpyxl not installed; skipping", file=sys.stderr)
        return

    import datetime
    import openpyxl

    zip_url = (
        "https://www.bankofengland.co.uk/-/media/boe/files/statistics/"
        "yield-curves/glcnominalddata.zip"
    )
    data = fetch(zip_url, timeout=300)
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xlsx_name = "GLC Nominal daily data_2025 to present.xlsx"
        with z.open(xlsx_name) as src:
            wb = openpyxl.load_workbook(io.BytesIO(src.read()), data_only=True)

    ws = wb["4. spot curve"]
    # Row 4 is the tenor header; column 1 ("Maturity years:") is prefix.
    header = [c.value for c in ws[4]]
    col_by_tenor = {t: header.index(t) for t in CURVE_TENORS if t in header}

    target = datetime.date(2025, 12, 31)
    row = next(
        r for r in ws.iter_rows(min_row=6, values_only=True)
        if isinstance(r[0], datetime.datetime) and r[0].date() == target
    )

    out = HERE / "uk_gilt_20251231.csv"
    with out.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["tenor_years", "rate_pct_continuous", "rate_pct_semi_annual"])
        for t in CURVE_TENORS:
            if t not in col_by_tenor:
                continue
            r_cont = row[col_by_tenor[t]]
            w.writerow([t, f"{r_cont:.6f}", f"{cont_to_sa(r_cont):.4f}"])
    print(f"wrote {out}")


def pull_ecb_aaa() -> None:
    """ECB euro-area AAA spot curve on 2025-12-31 (Svensson spot yields)."""
    tenors = "+".join(f"SR_{t}Y" for t in CURVE_TENORS)
    url = (
        "https://data-api.ecb.europa.eu/service/data/YC/"
        f"B.U2.EUR.4F.G_N_A.SV_C_YM.{tenors}"
        f"?startPeriod={VAL_DATE}&endPeriod={VAL_DATE}&format=csvdata"
    )
    raw = fetch(url).decode("utf-8", errors="replace")
    rdr = csv.DictReader(io.StringIO(raw))

    by_tenor: dict[int, float] = {}
    for r in rdr:
        # KEY ends in e.g. "SR_10Y"; map back to int years.
        tag = r["KEY"].split(".")[-1]
        if not tag.startswith("SR_") or not tag.endswith("Y"):
            continue
        t = int(tag[3:-1])
        by_tenor[t] = float(r["OBS_VALUE"])

    out = HERE / "ecb_aaa_20251231.csv"
    with out.open("w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["tenor_years", "rate_pct_continuous", "rate_pct_semi_annual"])
        for t in CURVE_TENORS:
            if t not in by_tenor:
                continue
            r_cont = by_tenor[t]
            w.writerow([t, f"{r_cont:.6f}", f"{cont_to_sa(r_cont):.4f}"])
    print(f"wrote {out}")


def pull_jgb() -> None:
    """MOF Japan JGB par yields on the last business day of 2025.

    JP markets close 12-31 for year-end; last observation is typically 12-30.
    CSV columns: Date, 1Y, 2Y, 3Y, 4Y, 5Y, 6Y, 7Y, 8Y, 9Y, 10Y, 15Y, 20Y,
    25Y, 30Y, 40Y. Values are par yields in percent, semi-annual market
    convention — stored as-is.
    """
    url = (
        "https://www.mof.go.jp/english/policy/jgbs/reference/interest_rate/"
        "historical/jgbcme_all.csv"
    )
    raw = fetch(url, timeout=120).decode("utf-8", errors="replace")
    lines = raw.splitlines()
    # Line 2 is the header: "Date,1Y,2Y,..."
    header = lines[1].split(",")
    col_by_tenor = {int(h.rstrip("Y")): i for i, h in enumerate(header)
                    if h.endswith("Y") and h[:-1].isdigit()}

    # Find the latest row in 2025-12; MOF uses YYYY/M/D (un-padded) format,
    # so parse to (y, m, d) before comparing — string order would rank
    # "2025/12/9" above "2025/12/30".
    import datetime
    latest_row = None
    latest_dt = datetime.date.min
    for ln in lines[2:]:
        first = ln.split(",", 1)[0]
        if not first.startswith("2025/12/"):
            continue
        y, m, d = (int(x) for x in first.split("/"))
        dt = datetime.date(y, m, d)
        if dt > latest_dt:
            latest_dt = dt
            latest_row = ln
    if latest_row is None:
        raise RuntimeError("no 2025-12 row found in MOF JGB CSV")
    latest_date = latest_dt.isoformat()

    cells = latest_row.split(",")
    out = HERE / f"jgb_{latest_date}.csv"
    # Also write a stable-name file for convenience.
    stable = HERE / "jgb_eoy2025.csv"
    for path in (out, stable):
        with path.open("w", newline="") as fh:
            w = csv.writer(fh)
            w.writerow(["tenor_years", "rate_pct"])
            for t in CURVE_TENORS:
                if t not in col_by_tenor:
                    continue
                w.writerow([t, cells[col_by_tenor[t]]])
    print(f"wrote {out} and {stable} (observation {latest_date})")


# ---------------------------------------------------------------------------- main

def main() -> int:
    failed = []
    for name, fn in [
        ("UST CMT", pull_ust_cmt),
        ("SOFR fixings", pull_sofr_fixings),
        ("TIPS index ratio", pull_tips_index_ratio),
        ("UK Gilt", pull_uk_gilt),
        ("ECB AAA", pull_ecb_aaa),
        ("JGB", pull_jgb),
    ]:
        try:
            fn()
        except Exception as e:  # noqa: BLE001
            print(f"{name} pull failed: {e}", file=sys.stderr)
            failed.append(name)
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
